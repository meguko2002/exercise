#! python3
# updateProduce.py - 農産物スプレッドシートの価格を訂正する

import openpyxl

wb = openpyxl.load_workbook('fileEdit/produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# 農産物の種類と、更新する価格
PRICE_UPDATES = {'Garlic': 33.01,
                'Celery': 1.19,
                'Lemon': 1.27}

# TODO: 行をループして価格を更新する
for row_num in range(2, sheet.max_row + 1): #先頭行をスキップ
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

wb.save('fileEdit/updateProduceSales.xlsx')
