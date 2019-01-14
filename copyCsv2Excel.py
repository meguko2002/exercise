#! python3
#removeCsvHeader.py

import csv, os
import openpyxl

#入力ファイル、出力ファイルの保管フォルダを作成する
os.makedirs('inputFiles', exist_ok=True)
os.makedirs('outputFiles', exist_ok=True)

#inputFilesフォルダの全ファイルをループする
for csv_filename in os.listdir('inputFiles'):
    if not csv_filename.endswith('.csv'):
        continue #CSVファイルでなければスキップ
    print('見出し削除中' + csv_filename + '...')

    # CSVファイルを読み込んで各行をcsv_rows[]に書き込む
    csv_rows = []
    csv_file_obj = open('fileEdit/' + csv_filename)
    reader_obj = csv.reader(csv_file_obj)
    for row in reader_obj:
        csv_rows.append(row)
    csv_file_obj.close()

    # 保存先のExceファイルを作成
    wb = openpyxl.load_workbook('ooutputFile/' + csv_filename)
    sheet = wb.get_sheet_by_name('Sheet')

    # CSVファイルから読み込んだcsv_rows[]をExcelファイルに書き出す
    for row_num in range(len(csv_rows)):  # 先頭行をスキップ
        produce_name = sheet.cell(row=row_num, column=1).value
        if produce_name in PRICE_UPDATES:
            sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

    wb.save('fileEdit/updateProduceSales.xlsx')


