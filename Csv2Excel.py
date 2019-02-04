#! python3

import csv, os, shutil, cv2
import openpyxl
import numpy as np

waveTypeArray = ('t', 'd', 'v', 'd')    # データタイプ (t:time, d:変位, v:速度, n:データなし)

# inputFilesフォルダの全ファイルをループする
for csv_filename in os.listdir('inputFiles'):
    if not csv_filename.endswith('.csv'):
        continue  # CSVファイルでなければスキップ
    print(csv_filename + ' 処理中')
    # CSVファイルを読み込んで各行をcsv_rows[]に書き込む
    csv_rows = []
    csv_file_obj = open('inputFiles/' + csv_filename)
    reader_obj = csv.reader(csv_file_obj)
    for row in reader_obj:
        csv_rows.append(row)
    csv_file_obj.close()

    # rowとcolumnを転置して波形ごとに要素を分ける
    dataset_t = np.array(csv_rows)
    waveset = dataset_t.T.tolist()  # 生波形を格納する
    waveset = [[float(x) for x in wave] for wave in waveset] # str -> float
    datanum = len(waveset[0])   # データ点数

    # dataset各列について処理
    wave_title = ['time[s]']    # 時間波形のタイトル
    dn, vn = 1, 1               # 波形の通し番号
    add_waveset = []            # 生波形から誤差率を計算した波形を格納する
    add_wave_title = []         # 誤差率処理後の時間波形のタイトル
    fft_title = ['f[Hz]']       # FFT波形のタイトル
    fftset = []               # FFT波形を格納する

    # サンプル数Ｎの決定
    n = 1
    while True:
        if 2 ** n > datanum:
            N = 2 ** (n-1)
            break
        else:
            n += 1

    # 周波数軸のデータ作成
    dt = waveset[0][1]-waveset[0][0]
    fq = np.linspace(0, 1.0/dt, N).tolist()
    fftset.append(fq)

    # 各生波形をデータ処理
    try:
        for raw_wave in waveset:
            if waveTypeArray[waveset.index(raw_wave)] == 'd':
                wave_title.append('変位' + str(dn))
                # 生波形をfft
                fft_title.append('変位' + str(dn))
                F = np.fft.fft(raw_wave)
                F_abs_amp = np.abs(F).tolist()
                fftset.append(F_abs_amp)
                dn += 1
            elif waveTypeArray[waveset.index(raw_wave)] == 'v':
                wave_title.append('周速' + str(vn))
                # 誤差率処理
                arr_wave = np.array(raw_wave)
                wave = ((arr_wave - np.mean(arr_wave))/np.mean(arr_wave)).tolist()
                add_waveset.append(wave)
                add_wave_title.append('周速' + str(vn) + ' 誤差')
                # 誤差率波形をfft
                fft_title.append('周速' + str(vn))
                F = np.fft.fft(wave)
                F_abs_amp = np.abs(F).tolist()
                fftset.append(F_abs_amp)
                vn += 1
            elif waveTypeArray[waveset.index(raw_wave)] == 'n':
                wave_title.append('データなし')

    except IndexError:
        print('waveTypeArray にデータタイプの定義が足りないかも')


    # ひな型ファイルの読み込み
    wb = openpyxl.load_workbook('hinagata.xlsx')
    # wavesetをシートtimeに書き出す
    sheet = wb['time']
    for col_num in range(len(wave_title)):
        sheet.cell(row=1, column=col_num+1, value=wave_title[col_num])
    for col_num in range(len(wave_title)):
        for row_num in range(datanum):
            sheet.cell(row=row_num+2, column=col_num+1, value=waveset[col_num][row_num])

    # fftsetをシートfftに書き出す
    sheet = wb['fft']
    for col_num in range(len(fft_title)):
        sheet.cell(row=1, column=col_num+1, value=fft_title[col_num])
    for col_num in range(len(fft_title)):
        for row_num in range(N):
            sheet.cell(row=row_num+2, column=col_num+1, value=fftset[col_num][row_num])

    wb.save('outputFiles/' + csv_filename.split('.')[0] + '.xlsx')


# # csvファイルと同名のtifファイルを読み込み
# for tif_filename in os.listdir('inputFiles'):
#     if tif_filename == csv_filename.split('.')[0] + '.tif':
#         img = cv2.imread('inputFiles/' + tif_filename, 0)
#         print(img)
#         # 各rowの要素を平均化して1次元化
#         img_ave = np.average(img, axis=1)
